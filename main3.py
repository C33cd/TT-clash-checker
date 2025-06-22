from openpyxl.reader.excel import load_workbook

if __name__ == '__main__':
    wb1 = load_workbook(filename='TIMETABLE - II SEMESTER 2024 -25_removed.xlsx')
    wb2 = load_workbook(filename='DRAFT TIMETABLE (1).xlsx')
    ws1 = wb1.active
    ws2 = wb2.active
    print(ws1.max_row)
    print(ws2.max_row)
    i = 1
    while True:
        if ws1.cell(row=i)