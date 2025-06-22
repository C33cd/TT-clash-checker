from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

_file_name = "Test.xlsx"
_sheet_name = "Test_Sheet"


def new_workbook(_file_name, _sheet_name):
    wb = Workbook()  # Workbook Object
    ws = wb.active  # Gets the active worksheet
    ws.title = _sheet_name  # Name the active worksheet

    # Writing the header columns
    ws['A1'] = 'Name'
    ws['B1'] = 'Class'
    ws['C1'] = 'Section'
    ws['D1'] = 'Marks'
    ws['E1'] = 'Age'

    col_range = ws.max_column  # get max columns in the worksheet

    # formatting the header columns, filling red color
    for col in range(1, col_range + 1):

        ws.cell(1, col).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                       fill_type="solid")  # used hex code for red color
        #print(cell_header.fill)
        #print()

    wb.save(_file_name)  # save the workbook
    wb.close()  # close the workbook


if __name__ == '__main__':
    new_workbook(_file_name, _sheet_name)

    """
                j=0
                l = '0'
                l_flag = False
                p = '0'
                p_flag = False
                t = '0'
                t_flag = False
                # do while loop -> executes at least once:

                while True:
                    if ws.cell(row=i+j,column=7).value[0]=='L':
                        temp = ws.cell(row=i + j, column=10).value + '' + ws.cell(row=i + j, column=11).value
                        if ws.cell(row=i + j + 1, column=7).value[0] is None:
                            temp += ws.cell(row=i + j + 1, column=10).value + '' + ws.cell(row=i + j + 1, column=11).value

                        if l =='0':
                            l = temp
                        elif temp!=l:
                            l = '0'
                            l_flag = True
                            break
                    elif ws.cell(row=i+j,column=7).value[0]=='P':
                        if p=='0':
                            p = ws.cell(row=i+j,column=10).value + '' + ws.cell(row=i+j,column=11).value
                        elif ws.cell(row=i+j,column=10).value + '' + ws.cell(row=i+j,column=11).value!=p:
                            p='0'
                            p_flag = True
                            break
                    elif ws.cell(row=i+j,column=7).value[0]=='T':
                        if t=='0':
                            t = ws.cell(row=i+j,column=10).value + '' + ws.cell(row=i+j,column=11).value
                        elif ws.cell(row=i+j,column=10).value + '' + ws.cell(row=i+j,column=11).value!=t:
                            t='0'
                            t_flag = True
                            break
                    j+=1
                    # conditional to end loop
                    if ws.cell(row=i + j, column=2).value is not None:
                        break
                if not l_flag:
                if not p_flag:
                if not t_flag:
                    for c in range(0, len(t)):
                        timetable[int(t[c])][Days[t[c]].value] = False
            """