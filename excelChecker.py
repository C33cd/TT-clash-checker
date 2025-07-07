# This is a sample Python script.
from typing import List

import openpyxl
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

from openpyxl.reader.excel import load_workbook
from enum import Enum

from openpyxl.styles import PatternFill


# Use a breakpoint in the code line below to debug your script.
# Press Ctrl+F8 to toggle the breakpoint.
class Days(Enum):
    M = 0
    T = 1
    W = 2
    Th = 3
    F = 4


class ExcelChecker:

    def store_as_tuple(self, wb1: openpyxl.workbook.workbook.Workbook, ind1: int, arr_sect: set) -> int:
        # arr_sect -> stores array of sections of particular type (L/P/T)
        ws = wb1.active  # same worksheet as main program
        section_hours = set()
        # storing format -> (Day, hour)
        # loop to collect and store all the  hours
        ind2 = 0  # j -> loop variable to traverse the rows
        if ws.cell(row=ind1 + ind2, column=10).fill!=PatternFill(start_color='FF0000', end_color='FF0000',
                                    fill_type="solid"):

            while True:
                s1 = str(ws.cell(row=ind1 + ind2, column=10).value).split(' ')
                if str(ws.cell(row=ind1 + ind2, column=10).value) != 'None' and str(
                        ws.cell(row=ind1 + ind2, column=10).value) != 'DAYS':
                    # traverse through s1, store hours in p as tuple
                    for ch in s1:
                        s2 = str(ws.cell(row=ind1 + ind2, column=11).value).split(' ')
                        # add tuple
                        for times in s2:
                            section_hours.add((Days[ch].value, int(times)))
                ind2 += 1
                # continue if there is nothing in column 7
                if ws.cell(row=ind1 + ind2, column=7).value is None:
                    pass
                else:
                    # decrement ind2 by 1 as it will be incremented by 1 outside the loop
                    ind2 -= 1
                    break

        # add each section individually to arr_sect
        # convert section_hours to a frozen set so that we can store it in a set(arr_sect).
        # We will be storing all the values at one time. Hence, using frozenset doesn't cause issues
        section_hours = frozenset(section_hours)
        if section_hours != set():
            arr_sect.add(section_hours)

        # the no. of lines to be skipped to go to next type of course is returned
        return ind2

    def check_for_clashes(self, wb1:openpyxl.workbook.workbook.Workbook, inc: int, i: int, colored_red: bool,
                          lectures:List[set], tuts: List[set], practicals: List[set], timetable:List[List[bool]]):
        # storing format -> (Day, hour)
        # loop to collect all the lecture hours
        ws = wb1.active
        j = inc
        row_head = i + inc
        l_hours = set()
        l_hours_times = set()
        while True:
            if str(ws.cell(row=i + j, column=10).value) != 'None' and str(
                    ws.cell(row=i + j, column=10).value) != 'DAYS':
                # get lecture dates and times
                lecture_days = str(ws.cell(row=i + j, column=10).value).split(' ')
                lecture_times = str(ws.cell(row=i + j, column=11).value).split(' ')
                print(i+j)
                print(lecture_times)
                # traverse through lecture_days, check if feasible to add
                for day in lecture_days:
                    for time in lecture_times:
                        l_hours.add(frozenset({(Days[day].value, int(time))}))
                        l_hours_times.add((Days[day].value, int(time)))
                        if timetable[int(time) - 1][Days[day].value]:
                            pass
                        else:
                            colored_red = True
                            # color row red
                            for col in range(1, 14):
                                ws.cell(row=row_head, column=col).fill = PatternFill(
                                    start_color='FF0000',
                                    end_color='FF0000',
                                    fill_type="solid")

            j += 1

            # end condition for do-while loop
            # end loop as soon as it reaches a new section
            if ws.cell(row=i + j, column=7).value is not None or i + j >= ws.max_row:
                if not colored_red:
                    for t1 in l_hours_times:
                        for ind in range(0, len(lectures)):
                            clash = 0
                            for j in lectures[ind]:
                                for k in j:
                                    if k == t1:
                                        clash += 1
                                        break
                            if clash>=len(lectures[ind]) and clash != 0:
                                # print('L: ' + str(ws.cell(row=row_head, column=2).value))
                                # color row red
                                for col in range(1, 14):

                                    ws.cell(row=row_head, column=col).fill = PatternFill(
                                            start_color='FF0000',
                                            end_color='FF0000',
                                            fill_type="solid")
                                break

                    for t1 in tuts:
                        if t1.issubset(l_hours) and t1 != set():
                            # print('T: ' + str(ws.cell(row=row_head, column=2).value))
                            # color row red
                            for col in range(1, 14):
                                ws.cell(row=row_head, column=col).fill = PatternFill(
                                    start_color='FF0000',
                                    end_color='FF0000',
                                    fill_type="solid")
                    for t1 in l_hours_times:
                        for ind in range(0, len(practicals)):
                            clash = 0
                            for j in practicals[ind]:
                                for k in j:
                                    if k == t1:
                                        clash += 1
                                        break
                            if clash>=len(practicals[ind]) and clash != 0:
                                # print('P: ' + str(ws.cell(row=row_head, column=2).value))
                                # color row red
                                for col in range(1, 14):

                                    ws.cell(row=row_head, column=col).fill = PatternFill(
                                            start_color='FF0000',
                                            end_color='FF0000',
                                            fill_type="solid")
                                break

                break

    # color code: Red = midsem/compre/class at same time as cdc, yellow = midsem/compre on same day as cdc
    # Assumption: CDCs are clash-free
    def mainchecker(self, cdcs: List[str], f_name: str):
        timetable = [[True for x in range(5)] for y in range(12)]
        wb = load_workbook(filename=f_name)
        ws = wb.active

        # Pre-processing:
        print('Pre-processing...')
        # code for altering file for accurate representation
        for i in range(3, ws.max_row + 1):
            # check the hours column and format accordingly:
            if ws.cell(row=i, column=11).value is not None:
                val = str(ws.cell(row=i, column=11).value)
                if len(val) == 4:
                    if val[0]=='1':
                        if val[1]<=val[0]:
                            ws.cell(row=i, column=11).value = val[0] + val[1] + ' ' + val[2] + val[3]
                        else:
                            ws.cell(row=i, column=11).value = val[0] + ' ' + val[1] + ' ' + val[2] + val[3]
                    elif val[2]=='1':
                        ws.cell(row=i, column=11).value = val[0] + ' ' + val[1] + ' ' + val[2] + val[3]
                    else:
                        ws.cell(row=i, column=11).value = val[0] + val[1] + ' ' + val[2] + val[3]
                elif len(val) == 3:
                    if val[1] == '1':
                        ws.cell(row=i, column=11).value = val[0] + ' ' + val[1] + val[2]
                    elif val[0] == '1':
                        ws.cell(row=i, column=11).value = val[0] + val[1] + ' ' + val[2]
                elif len(val) == 2 and val != '10':
                    ws.cell(row=i, column=11).value = val[0] + ' ' + val[1]
            # if ws.cell(row=i, column=1).value=='COMP CODE' or str(ws.cell(row=i, column=1).value).startswith('DRAFT TIMETABLE'):
            # delete row
        wb.save(filename='DRAFT TIMETABLE (2).xlsx')
        wb.close()

        # code to work on new file
        wb = load_workbook(filename='DRAFT TIMETABLE (2).xlsx')
        ws = wb.active

        cdcs.sort()  # so that midsem_cdcs and compre_cdcs dates match index-wise with cdcs subject
        # note that Excel sheet has courses sorted in alphabetical order of course codes

        # arrays to store midsem and compre dates of cdcs
        midsem_cdcs = []
        compre_cdcs = []

        # arrays to store lectures, practicals and tut timings of cdcs
        lectures = []
        practicals = []
        tuts = []

        # storing values of cdcs by traversing through the file
        print('Storing CDC data...')
        count = 0  # stores no. of cdcs counted -> eliminates the need to traverse entire doc
        i = 1
        while i <= ws.max_row:
            if ws.cell(row=i, column=2).value in cdcs:
                count += 1
                # store midsem and compre dates if cdc has a midsem and compre
                if ws.cell(row=i, column=12).value is not None:
                    midsem_cdcs.append(ws.cell(row=i, column=12).value)
                if ws.cell(row=i, column=13).value is not None:
                    compre_cdcs.append(ws.cell(row=i, column=13).value)
                # store unique slots and timings of classes:
                l_sect = set()  # stores individual sections
                p_sect = set()  # stores individual sections
                t_sect = set()  # stores individual sections

                # do-while loop to store values
                # lectures stored as set of (sets of lecture hours)
                # tutorials and practicals stored as set of timings
                j = 0
                while True:
                    # in case of program encountering one of the blank spaces(tuts/lectures with 1+ instructors)
                    if ws.cell(row=i + j, column=7).value is None:
                        pass
                    # code for storing practicals -> same format follows for lectures and tuts as well
                    elif ws.cell(row=i + j, column=7).value[0] == 'P':
                        # storing format -> (Day,Hour) in frozenset
                        j += self.store_as_tuple(wb, i + j, p_sect)

                    # code for storing tuts
                    elif ws.cell(row=i + j, column=7).value[0] == 'T':
                        j += self.store_as_tuple(wb, i + j, t_sect)

                    # code for storing lectures
                    elif ws.cell(row=i + j, column=7).value[0] == 'L':
                        j += self.store_as_tuple(wb, i + j, l_sect)

                    j += 1  # this has to be outside the if-statements to ensure loop keeps moving forward
                    # condition to break out of loop
                    if ws.cell(row=i + j, column=2).value is not None and ws.cell(row=i + j,
                                                                                  column=2).value != 'COURSE NO.':
                        break
                # add timings to arrays
                lectures.append(l_sect)
                practicals.append(p_sect)
                tuts.append(t_sect)
                i += j  # takes i to 'row number' of next course
            # stop traversing if all the cdcs are covered
            elif count == len(cdcs):
                break
            else:
                i += 1  # moves i to next row

        # print unique lectures, tutorials and practicals -> commented out
        """
        print('[')
        for el in lectures:
            print(el)
        print(']')
        print('[')
        for el in tuts:
            print(el)
        print(']')
        print('[')
        for el in practicals:
            print(el)
        print(']')
        """

        # code to allot confirmed cdc slots in timetable
        print('Confirming timetable...')

        # for tuts -> if no of elements in set = no of duplicates of set, then confirm all the timings in that set
        for i in range(0, len(tuts)):
            dupl = 0  # stores no. of duplicates
            for j in range(0, len(tuts)):
                if tuts[i] == tuts[j]:
                    dupl += 1
            if dupl == len(tuts[i]):
                # confirm lecture slots in timetable using logic given above(previous comment)
                for section in tuts[i]:
                    for timing in section:
                        # allot slots in timetable, timing[1]-1 is because timing[1] has lecture hour which starts from 1
                        # if timetable[timing[1]-1][timing[0]] is true, that slot is free
                        # if it is false, that spot has already been taken
                        if timetable[timing[1] - 1][timing[0]]:
                            timetable[timing[1] - 1][timing[0]] = False
                        else:
                            print('CDC clash at: ' + str(timing))
                # remove all subsets of confirmed lecture timings
                # (if lecture timings are confirmed, you can't take another class during that time)
                c1 = 0
                while c1 < len(tuts):
                    tuts[c1] = tuts[c1] - tuts[i]
                    c1 += 1
            elif dupl > tuts[i].__len__() and tuts[i] != set():
                print('CDC clash at timings: ' + str(tuts[i]))

        # same logic as tuts for confirming lectures in timetable
        for i in range(0, len(lectures)):
            dupl = 0  # stores no. of duplicates
            for j in range(0, len(lectures)):
                if lectures[i] == lectures[j]:
                    dupl += 1
            if dupl == len(lectures[i]):
                # confirm lecture slots in timetable using logic given above(previous comment)
                for section in lectures[i]:
                    for timing in section:
                        # allot slots in timetable, timing[1]-1 is because timing[1] has lecture hour which starts from 1
                        # if timetable[timing[1]-1][timing[0]] is true, that slot is free
                        # if it is false, that spot has already been taken
                        if timetable[timing[1] - 1][timing[0]]:
                            timetable[timing[1] - 1][timing[0]] = False
                        else:
                            print('CDC clash at: ' + str(timing))
                # remove all subsets of confirmed lecture timings
                # (if lecture timings are confirmed, you can't take another class during that time)
                c1 = 0
                while c1 < len(lectures):
                    lectures[c1] = lectures[c1] - lectures[i]
                    c1 += 1
            elif dupl > lectures[i].__len__() and lectures[i] != set():
                print('CDC clash at timings: ' + str(lectures[i]))

        # code for storing practicals -> logic same as lectures
        for i in range(0, len(practicals)):
            dupl = 0  # stores no. of duplicates
            for j in range(0, len(practicals)):
                if practicals[i] == practicals[j]:
                    dupl += 1
            if dupl == len(practicals[i]):
                # confirm lecture slots in timetable using logic given above(previous comment)
                for section in practicals[i]:
                    for timing in section:
                        # allot slots in timetable, timing[1]-1 is because timing[1] has practical hour which starts from 1
                        # if timetable[timing[1]-1][timing[0]] is true, that slot is free
                        # if it is false, that slot has already been taken
                        if timetable[timing[1] - 1][timing[0]]:
                            timetable[timing[1] - 1][timing[0]] = False
                        else:
                            print('CDC clash at: ' + str(timing))
                # remove all subsets of confirmed lecture timings
                # (if lecture timings are confirmed, you can't take another class during that time)
                c1 = 0
                while c1 < len(practicals):
                    practicals[c1] = practicals[c1] - practicals[i]
                    c1 += 1
            elif dupl > practicals[i].__len__() and practicals[i] != set():
                print('CDC clash at timings: ' + str(practicals[i]))

        # printing confirmed data ->commented out
        """
        print()
        print('Confirmed data: ')
        print(timetable)
        print('[')
        for el in lectures:
            print(el)
        print(']')
        print('[')
        for el in tuts:
            print(el)
        print(']')
        print('[')
        for el in practicals:
            print(el)
        print(']')
        print()
        """

        # code to check clashes
        print()
        print('Checking clashes :')
        print()

        for i in range(1, ws.max_row + 1):
            # print statement to show script is working
            if i % 500 == 0:
                print('Processing...')
            if ws.cell(row=i, column=2).value is not None:
                colored_red = False
                for j in range(0, len(midsem_cdcs)):
                    # check for direct clash b/w midsem/compre -> same day, same time
                    if midsem_cdcs[j] == ws.cell(row=i, column=12).value or compre_cdcs[j] == ws.cell(row=i,
                                                                                                      column=13).value:
                        inc = 0  # increments upto next course
                        colored_red = True  # states that the course has been colored red
                        # write midsem/compre clash in adjacent cell
                        if midsem_cdcs[j] == ws.cell(row=i, column=12).value:
                            ws.cell(row=i + inc, column=14).value = 'Midsem clash with: ' + cdcs[j]
                        else:
                            ws.cell(row=i + inc, column=14).value = 'Compre clash with: ' + cdcs[j]
                        # color the entire course red
                        while True:
                            for col in range(1, 14):
                                # set fill of cells to red
                                ws.cell(row=i + inc, column=col).fill = PatternFill(start_color='FF0000',
                                                                                    end_color='FF0000',
                                                                                    fill_type="solid")
                            inc += 1
                            if (ws.cell(row=i + inc, column=2).value is not None and
                                    ws.cell(row=i + inc, column=2).value != 'COURSE NO.') or \
                                    i + inc == ws.max_row:
                                break
                    # below code handles NoneType subscript-able error for next elif
                    elif ws.cell(row=i, column=12).value is None or ws.cell(row=i, column=13).value is None:
                        break
                    # check if midsem/compre on same day
                    elif midsem_cdcs[j][:5] == ws.cell(row=i, column=12).value[:5] or \
                            compre_cdcs[j][:5] == ws.cell(row=i, column=13).value[:5]:

                        inc = 0  # increments upto next course

                        # write midsem/compre on same day
                        if midsem_cdcs[j][:5] == ws.cell(row=i, column=12).value[:5]:
                            ws.cell(row=i + inc, column=14).value = 'Midsem on same day as: ' + cdcs[j]
                        else:
                            ws.cell(row=i + inc, column=14).value = 'Compre on same day as: ' + cdcs[j]

                        # color full course yellow:
                        while True:
                            for col in range(1, 14):
                                # setting fill to yellow:
                                ws.cell(row=i + inc, column=col).fill = PatternFill(start_color='FFFF00',
                                                                                    end_color='FFFF00',
                                                                                    fill_type="solid")
                            inc += 1
                            if (ws.cell(row=i + inc, column=2).value is not None and
                                    ws.cell(row=i + inc, column=2).value != 'COURSE NO.') or i + inc == ws.max_row:
                                break

                # check for class clashes (among unconfirmed classes) if no midsem/compre clashes
                if not colored_red:
                    inc = 0  # increments upto next course
                    while True:
                        if ws.cell(row=i + inc, column=7).value is None:
                            pass
                        # code for checking practicals, tuts and lectures
                        else:
                            self.check_for_clashes(wb, inc, i, colored_red, lectures, tuts, practicals, timetable)

                        inc += 1

                        # end condition for do-while loop:
                        if ws.cell(row=i + inc, column=2).value is not None or (i + inc) >= ws.max_row:
                            break

        # save and close the file
        wb.save(filename='DRAFT TIMETABLE (2).xlsx')
        wb.close()
        print("Done, please check the excel sheet DRAFT TIMETABLE(2)")
